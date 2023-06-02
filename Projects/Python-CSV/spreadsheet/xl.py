import openpyxl as xl
from openpyxl.chart import BarChart, Reference

command = input("Filename: ")

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for each_row in range(2,sheet.max_row + 1):
        cell = sheet.cell(each_row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(each_row,4)
        corrected_price_cell.value = corrected_price
        print(cell.value)

    values = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save("transactions2.xlsx")

output = process_workbook(command)
